import speech_recognition as sr
import time
import os
from openpyxl import  load_workbook

#Funcao responsavel por ouvir e reconhecer a fala
def ouvir_microfone(microfone):
	with sr.Microphone() as source:
		#Chama a funcao de reducao de ruido disponivel na speech_recognition
		microfone.adjust_for_ambient_noise(source)
		#Armazena a informacao de audio na variavel
		print("FALE")
		audio = microfone.listen(source)
		return audio

def converte_audio(audio):
	try:
		# Passa o audio para o reconhecedor de padroes do speech_recognition
		frase = microfone.recognize_google(audio, language='pt-BR')
		# Após alguns segundos, retorna a frase falada
		# print("Você disse: " + frase)
		return frase
	# Caso nao tenha reconhecido o padrao de fala, exibe esta mensagem
	except:
		frase = "faild"
		return frase


def posicao():
	# Abre o arquivo para Leitura
	with open("infor.txt", "r", encoding="utf-8") as arquivo:
		posicao = arquivo.read()
		arquivo.close()
	return posicao


def novaposicao(posicao):
	cont = int(posicao) + 1
	texto = str(cont)
	# Abre o arquivo para escrita
	with open("infor.txt", "w", encoding="utf-8") as arquivo:
		arquivo.write(texto)


def cria_pasta(nome_escrito):
	# CRIA UMA PASTA COM O NOME DO USUARIO
	if not os.path.exists('audios/' + nome_escrito):
		os.makedirs('audios/' + nome_escrito)
	else:
		#Caso tenha uma pasta com o mesmo nome
		pos = posicao()
		nome_escrito = nome_escrito + pos
		os.makedirs('audios/' + nome_escrito)
	return nome_escrito


def gravadados(coluna, linha , dados):
	wb = load_workbook('dados/Arquivo.xlsx')
	# Use a aba ativa quando o arquivo foi carregado
	ws = wb.active
	ws[coluna+linha] = dados
	# Salve
	wb.save('dados/Arquivo.xlsx')

def salvaaudio(nome_escrito,tipo,audio):
	with open('audios/' + nome_escrito + "/"+ tipo+".wav", 'wb') as f:
		f.write(audio.get_wav_data())

if __name__ == '__main__':
	continue_voto = 1

	while continue_voto == 1:
		print("Teste de Reconhecimento de Números pela Voz !!!")
		print("Comece a falar quando aparecer à palavra \"FALE\"")
		# Habilita o microfone para ouvir o usuario
		microfone = sr.Recognizer()

		confirma = 'nao'
		nome_escrito = " "
		while confirma != 'sim':
			print("Fale seu Primeiro Nome ")
			nome = ouvir_microfone(microfone)
			nome_escrito = converte_audio(nome)
			print("Seu nome é : ", nome_escrito)
			print("Fale SIM casso esteja correto")
			print("Fale NÃO para corrigir")
			ver = ouvir_microfone(microfone)
			confirma = converte_audio(ver)


		#Cria a Pasta com o Nome do Usuario
		nome_escrito = cria_pasta(nome_escrito)
		#Salva o Audio do nome
		salvaaudio(nome_escrito,'Nome',nome)

		linha = posicao()
		gravadados('A',linha,nome_escrito)


		#Grava Idade
		print("Fale sua idade ")
		idade = ouvir_microfone(microfone)
		idade_escrito = converte_audio(idade)
		# Salva o Audio da idade
		salvaaudio(nome_escrito, 'Idade',idade)
		gravadados('B', linha, idade_escrito)




		# Grava Escolaridade
		print("Fale sua Escolaridade :  ")
		print("			Fundamental ")
		print("			Medio ")
		print("			Superior ")
		escolaridade = ouvir_microfone(microfone)
		escolaridade_escrito = converte_audio(escolaridade)
		# Salva o Audio da escolaridade
		salvaaudio(nome_escrito,'Escolaridade',escolaridade)
		gravadados('C', linha, escolaridade_escrito)


		print("Agora você só ira falar os números que aparecerem: ")
		time.sleep(1)
		print("123")
		num123 = ouvir_microfone(microfone)
		num123_escrito = converte_audio(num123)
		# Salva o Audio do numero123
		salvaaudio(nome_escrito, 'num123', num123)
		gravadados('D', linha, num123_escrito)
		time.sleep(2)

		print("1111")
		num1111 = ouvir_microfone(microfone)
		num1111_escrito = converte_audio(num1111)
		# Salva o Audio do numero1111
		salvaaudio(nome_escrito, 'num1111', num1111)
		gravadados('E', linha, num1111_escrito)
		time.sleep(2)

		print("90489")
		num90489 = ouvir_microfone(microfone)
		num90489_escrito = converte_audio(num90489)
		# Salva o Audio do num90489
		salvaaudio(nome_escrito, 'num90489', num90489)
		gravadados('F', linha, num90489_escrito)
		time.sleep(2)

		print("1")
		num1 = ouvir_microfone(microfone)
		num1_escrito = converte_audio(num1)
		# Salva o Audio do num1
		salvaaudio(nome_escrito, 'num1', num1)
		gravadados('G', linha, num1_escrito)

		print("2")
		num2 = ouvir_microfone(microfone)
		num2_escrito = converte_audio(num2)
		# Salva o Audio do num2
		salvaaudio(nome_escrito, 'num2', num2)
		gravadados('H', linha, num2_escrito)

		print("3")
		num3 = ouvir_microfone(microfone)
		num3_escrito = converte_audio(num3)
		# Salva o Audio do num3
		salvaaudio(nome_escrito, 'num3', num3)
		gravadados('I', linha, num3_escrito)


		print("4")
		num4 = ouvir_microfone(microfone)
		num4_escrito = converte_audio(num4)
		# Salva o Audio do num4
		salvaaudio(nome_escrito, 'num4', num4)
		gravadados('J', linha, num4_escrito)


		print("5")
		num5 = ouvir_microfone(microfone)
		num5_escrito = converte_audio(num5)
		# Salva o Audio do num5
		salvaaudio(nome_escrito, 'num5', num5)
		gravadados('K', linha, num5_escrito)


		novaposicao(linha)
		print("Obrigado por ter realizado o teste.")
		print("******* FIM *******")

		time.sleep(4)
		continue_voto = int(input("Continuar a Pesquisa "))


